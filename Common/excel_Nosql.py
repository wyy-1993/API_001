#! /usr/bin/env python
# -*- coding:utf-8 -*-

"""
@author:简凡
@file: excel_Nosql.py
@time: 2020/02/05  10:26
"""

import xlrd
import xlwt
from openpyxl import load_workbook
from Conf import testfile_path
from xlutils import copy

"""
主要是用于excel表格的读写：
* 读主要是用于读取测试用例数据；
* 写主要是将测试结果写入结果文件，
-- 需要注意，结果文件需要先自行创建，否正不会保存，因为目前没有判断文件是否存在。
"""

class Noexcel:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def find_row(self, regular):
        pass

    def write_regular(self, row, regular_result):
        # 将表达式提取的对应值，填写到表格中
        file_name = load_workbook(self.file_path)
        sheet = file_name[self.sheet_name]
        sheet.cell(row, 8).value = regular_result
        file_name.save(self.file_path)

    def write_data(self, row, run_code, run_result, test_result):
        # 将测试结果对应的状态码、接口返回结果、测试结果记录到表格
        b = load_workbook(self.file_path)
        sheet = b[self.sheet_name]
        sheet.cell(row, 11).value = run_code
        sheet.cell(row, 12).value = run_result
        sheet.cell(row, 13).value = test_result
        b.save(self.file_path)

    def Write_pass(self, url, method, param, run_result, test_result):
        dataexcel = xlrd.open_workbook(self.file_path)  # 打开测试结果保存的文件
        row = dataexcel.sheets()[self.sheet_name].nrows  # 读取当前文件的行数
        excel = copy.copy(dataexcel)
        table = excel.get_sheet(0)
        # row = rows
        styleBlueBkg1 = xlwt.easyxf('font: color-index red, bold on')
        table.write(row, 0, url)
        table.write(row, 1, method)
        table.write(row, 2, param)
        table.write(row, 3, run_result)
        table.write(row, 4, test_result, styleBlueBkg1)
        row += 1
        excel.save(self.file_path)

    def Write_fail(self, url, method, param, actual, result):
        rexcel = xlrd.open_workbook(self.file_path)
        rows = rexcel.sheets()[self.sheet_name].nrows
        excel = copy.copy(rexcel)
        table = excel.get_sheet(1)
        row = rows
        styleBlueBkg1 = xlwt.easyxf('font: color-index red, bold on')
        table.write(row, 0, url)
        table.write(row, 1, method)
        table.write(row, 2, param)
        table.write(row, 3, actual)
        table.write(row, 4, result, styleBlueBkg1)
        row += 1
        excel.save(self.file_path)


    def read_file(self):
        filename = load_workbook(self.file_path)  # 读取表名
        sheet = filename[self.sheet_name]  # 读取表单
        case_list = []

        for i in range(3, sheet.max_row + 1):

            test_data = {}
            test_data['case_id'] = sheet.cell(i, 1).value
            test_data['case_name'] = sheet.cell(i, 3).value
            test_data['method'] = sheet.cell(i, 4).value
            test_data['url'] = sheet.cell(i, 5).value
            test_data['param'] = sheet.cell(i, 6).value
            test_data['regular'] = sheet.cell(i, 7).value
            test_data['regular_result'] = sheet.cell(i, 8).value
            test_data['assert_content'] = sheet.cell(i, 9).value
            test_data['assert_code'] = sheet.cell(i, 10).value
            case_list.append(test_data)
        return case_list


if __name__ == '__main__':
    file_path = testfile_path.test_data_path
    regular_value = Noexcel(file_path,"test_data")

    case_list = Noexcel(file_path, "test_data").read_file()
    if case_list[1]["regular"] != None:
        regular_value.write_regular(3, "ceshinew")
        print("yizhixing")
    print(case_list[1])  # 返回修改前的数据
    ncase_list = Noexcel(file_path, "test_data").read_file()
    print(ncase_list[1])  # 返回修改后的
